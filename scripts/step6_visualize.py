#!/usr/bin/env python3
"""
step6_visualize.py

Generates STATIC summary charts and a formatted Excel report.

Behavior:
- Always generates baseline figures from outputs/step5/step5_eligibility_wide.csv

Final Updates:
- BAR CHART: Displays "Pass Rate %" above each bar instead of Total Count.
- EXCEL: Strict sorting (Pass -> Unclear No Abstract -> Unclear -> Fail).
- DATA: Pulls abstracts from step4_abstracts.csv.
- LAYOUT: Fixed absolute positioning for Titles/Legends.
"""

import re
import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.colors import ListedColormap
from matplotlib.patches import Patch
from openpyxl.styles import PatternFill, Alignment
import config as cfg

PALETTE = {
    "pass": "#d4edda",
    "fail": "#f8d7da",
    "unclear": "#fff3cd",
    "excel_pass": "C6EFCE",
    "excel_fail": "FFC7CE",
    "excel_unclear": "FFEB9C",
}


def _clean_doi(val):
    val = str(val).lower().strip()
    for p in ["https://doi.org/", "http://doi.org/", "doi:"]:
        if val.startswith(p):
            val = val[len(p) :]
    return val


def _get_sort_rank(val):
    val = str(val).lower().strip()
    if any(x in val for x in ["yes", "include", "pass"]):
        return 1
    if "no_abstract" in val:
        return 2
    if any(x in val for x in ["unclear", "pending"]):
        return 3
    if any(x in val for x in ["no", "exclude", "fail"]):
        return 4
    return 5


def _normalize_decision(val):
    val = str(val).lower().strip()
    if any(x in val for x in ["yes", "include", "pass"]):
        return 2
    if any(x in val for x in ["no", "exclude", "fail"]):
        return 0
    return 1


def _get_smart_key(row):
    doi = _clean_doi(row.get("doi", ""))
    if doi and doi not in ["nan", "none", ""]:
        return doi
    title = str(row.get("title", "")).strip().lower()
    return "".join(e for e in title if e.isalnum())


def _find_col(df, candidates):
    for c in candidates:
        if c in df.columns:
            return c
    return None


def _decision_bucket(v):
    s = str(v).strip().lower()
    if re.fullmatch(r"(yes|include|pass)", s):
        return "pass"
    if re.fullmatch(r"(no|exclude|fail)", s):
        return "fail"
    if "no_abstract" in s:
        return "no_abstract"
    if any(x in s for x in ["unclear", "pending", "skipped", "error"]):
        return "unclear"
    return "unclear"


def _load_abstract_map(step4_csv):
    abstract_map = {}
    if not os.path.exists(step4_csv):
        return abstract_map

    df_abs = pd.read_csv(step4_csv, engine="python", on_bad_lines="skip")
    abs_doi_col = _find_col(df_abs, ["doi", "DOI"])
    abs_txt_col = _find_col(df_abs, ["abstract", "Abstract", "text", "content"])

    if abs_doi_col and abs_txt_col:
        for _, row in df_abs.iterrows():
            d = _clean_doi(row[abs_doi_col])
            t = str(row[abs_txt_col])
            if d and t and t.lower() != "nan":
                abstract_map[d] = t
    return abstract_map


def _prepare_dataset(df_res, df_meta, abstract_map):
    t_col = _find_col(df_meta, ["title", "Title", "Study", "Article Title"])
    if t_col:
        df_meta = df_meta.rename(columns={t_col: "title"})
    else:
        df_meta["title"] = df_meta.get("doi", "")

    if "year" in df_meta.columns:
        df_meta = df_meta.drop(columns=["year"], errors="ignore")
    if "year_source" in df_meta.columns:
        df_meta = df_meta.drop(columns=["year_source"], errors="ignore")

    df_res = df_res.copy()
    df_meta = df_meta.copy()

    df_res["_join_key"] = df_res.apply(_get_smart_key, axis=1)
    df_meta["_join_key"] = df_meta.apply(_get_smart_key, axis=1)

    merged = pd.merge(df_res, df_meta[["_join_key", "title"]], on="_join_key", how="left")
    merged = merged.drop_duplicates(subset=["_join_key"])

    if "title_x" in merged.columns:
        merged.rename(columns={"title_x": "title"}, inplace=True)
    if "title_y" in merged.columns:
        merged["title"] = merged["title"].fillna(merged["title_y"])
    merged["title"] = merged["title"].fillna(merged.get("doi", ""))
    merged.drop(columns=["_join_key", "title_y"], inplace=True, errors="ignore")

    def get_abstract(row):
        d = _clean_doi(row.get("doi", ""))
        return abstract_map.get(d, "")

    merged["abstract"] = merged.apply(get_abstract, axis=1)
    merged["abstract"] = merged["abstract"].replace("", "No Abstract Available")

    crit_cols = sorted([c for c in merged.columns if "_decision" in c and "final" not in c.lower()])
    cols_to_fix = crit_cols + ["final_decision"]
    for c in cols_to_fix:
        if c not in merged.columns:
            merged[c] = "unclear"
        merged[c] = merged[c].astype(str)
        merged[c] = merged[c].replace([r"^\s*$", "nan", "None", "NaN"], "unclear", regex=True)
        merged[c] = merged[c].fillna("unclear")

    # Ensure notes column exists (do NOT include in decision columns)
    if "notes" not in merged.columns:
        merged["notes"] = ""
    merged["notes"] = merged["notes"].fillna("").astype(str)

    # Reorder: put notes immediately after final_decision (if present)
    if "final_decision" in merged.columns and "notes" in merged.columns:
        cols = list(merged.columns)
        cols.remove("notes")
        fi = cols.index("final_decision")
        cols.insert(fi + 1, "notes")
        merged = merged[cols]

    return merged, crit_cols, cols_to_fix


def _setup_fixed_header_layout(fig, height_inches, title_text, legend_elements, custom_top_margin=3.0):
    TITLE_Y_OFFSET = 0.5
    LEGEND_Y_OFFSET = 1.0

    top_frac = 1.0 - (custom_top_margin / height_inches)
    bottom_frac = 1.0 / height_inches if height_inches > 10 else 0.1

    plt.subplots_adjust(top=top_frac, bottom=bottom_frac, left=0.15, right=0.95)

    title_y = 1.0 - (TITLE_Y_OFFSET / height_inches)
    fig.suptitle(title_text, fontsize=18, weight="bold", y=title_y)

    legend_y = 1.0 - (LEGEND_Y_OFFSET / height_inches)
    fig.legend(
        handles=legend_elements,
        loc="upper center",
        bbox_to_anchor=(0.5, legend_y),
        ncol=3,
        frameon=False,
        fontsize=12,
    )


def _save_bar_summary(merged, cols_to_fix, viz_dir, filename, title_prefix):
    summary_data = []
    total_papers = len(merged)

    for col in cols_to_fix:
        counts = merged[col].value_counts()
        yes = counts[counts.index.str.contains("yes|include|pass", case=False, regex=True)].sum()
        no = counts[counts.index.str.contains("no|exclude|fail", case=False, regex=True)].sum()
        unc = total_papers - (yes + no)
        label = col.replace("_decision", "").replace("final_decision", "FINAL").upper()
        summary_data.append({"Criterion": label, "Pass": yes, "Fail": no, "Unclear": unc})

    df_sum = pd.DataFrame(summary_data).set_index("Criterion")[["Pass", "Fail", "Unclear"]]

    fig_height = 10
    fig, ax = plt.subplots(figsize=(12, fig_height))
    colors = [PALETTE["pass"], PALETTE["fail"], PALETTE["unclear"]]
    df_sum.plot(kind="bar", stacked=True, color=colors, edgecolor="black", width=0.7, ax=ax, legend=False)

    for c in ax.containers:
        labels = [int(v) if v > 0 else "" for v in c.datavalues]
        ax.bar_label(c, labels=labels, label_type="center", fontsize=10, color="black", weight="bold")

    totals = df_sum.sum(axis=1)
    pass_counts = df_sum["Pass"].values
    for i, total in enumerate(totals):
        p_count = pass_counts[i]
        p_rate = (p_count / total) * 100 if total > 0 else 0
        ax.text(i, total + 2, f"{p_rate:.1f}%", ha="center", weight="bold", fontsize=11, color="black")

    legend_elements = [
        Patch(facecolor=PALETTE["pass"], edgecolor="gray", label="Pass"),
        Patch(facecolor=PALETTE["fail"], edgecolor="gray", label="Fail"),
        Patch(facecolor=PALETTE["unclear"], edgecolor="gray", label="Unclear"),
    ]

    _setup_fixed_header_layout(fig, fig_height, f"{title_prefix} (N={total_papers})", legend_elements, custom_top_margin=2.0)

    plt.xticks(rotation=45, ha="right")
    ax.set_xlabel("Eligibility Criteria", labelpad=12)
    ax.set_ylabel("Number of Papers")
    plt.subplots_adjust(bottom=0.25)

    plt.savefig(os.path.join(viz_dir, filename), dpi=300)
    plt.close()


def _generate_long_heatmap(sub_df, crit_cols, viz_dir, title_text, filename):
    if sub_df.empty:
        return

    # --- replace applymap with per-column map (pandas 3+ compatible) ---
    temp_scores = sub_df[crit_cols].apply(lambda col: col.map(_normalize_decision))

    sub_df = sub_df.copy()
    sub_df["_sort_score"] = temp_scores.sum(axis=1)
    sub_df = sub_df.sort_values(by=["_sort_score", "title"], ascending=[False, True])

    # --- replace applymap again here ---
    matrix = sub_df[crit_cols].apply(lambda col: col.map(_normalize_decision))
    matrix.columns = [c.replace("_decision", "").upper() for c in crit_cols]
    matrix.index = sub_df["title"].astype(str).apply(lambda x: x[:80] + "..." if len(x) > 80 else x)

    n_rows = len(sub_df)
    fig_height = max(8, 4.0 + (n_rows * 0.35))

    fig = plt.figure(figsize=(12, fig_height))
    cmap = ListedColormap([PALETTE["fail"], PALETTE["unclear"], PALETTE["pass"]])
    ax = sns.heatmap(matrix, cmap=cmap, cbar=False, linewidths=0.5, linecolor="white", square=False, vmin=0, vmax=2)

    ax.xaxis.tick_top()
    ax.xaxis.set_label_position("top")
    plt.xticks(rotation=45, ha="left")

    legend_elements = [
        Patch(facecolor=PALETTE["pass"], edgecolor="gray", label="Pass"),
        Patch(facecolor=PALETTE["unclear"], edgecolor="gray", label="Unclear"),
        Patch(facecolor=PALETTE["fail"], edgecolor="gray", label="Fail"),
    ]

    _setup_fixed_header_layout(fig, fig_height, f"{title_text} ({len(sub_df)} Papers)", legend_elements, custom_top_margin=3.5)

    plt.savefig(os.path.join(viz_dir, filename), dpi=300)
    plt.close()

def _write_excel(final_df, crit_cols, out_excel):
    fill_pass = PatternFill(start_color=PALETTE["excel_pass"], end_color=PALETTE["excel_pass"], fill_type="solid")
    fill_fail = PatternFill(start_color=PALETTE["excel_fail"], end_color=PALETTE["excel_fail"], fill_type="solid")
    fill_unclear = PatternFill(start_color=PALETTE["excel_unclear"], end_color=PALETTE["excel_unclear"], fill_type="solid")
    align_single = Alignment(wrap_text=False, vertical="top")

    final_df = final_df.copy()
    final_df["_rank"] = final_df["final_decision"].apply(_get_sort_rank)
    final_df = final_df.sort_values(by=["_rank", "title"])
    final_df.drop(columns=["_rank"], inplace=True)

    cols = list(final_df.columns)
    if "abstract" in cols:
        cols.remove("abstract")
        cols.append("abstract")
    final_df = final_df[cols]

    with pd.ExcelWriter(out_excel, engine="openpyxl") as writer:
        final_df.to_excel(writer, sheet_name="All Papers", index=False)
        worksheet = writer.sheets["All Papers"]

        headers = [c.value for c in worksheet[1]]
        decision_headers = set(crit_cols + ["final_decision"])

        def classify(v):
            v = str(v).strip().lower()
            if re.fullmatch(r"(yes|include|pass)", v):
                return "pass"
            if re.fullmatch(r"(no|exclude|fail)", v):
                return "fail"
            if re.fullmatch(r"(unclear|pending|no_abstract)", v):
                return "unclear"
            return None

        for row in worksheet.iter_rows(
            min_row=2, max_row=len(final_df) + 1, min_col=1, max_col=len(cols)
        ):
            for cell in row:
                cell.alignment = align_single
                header = headers[cell.col_idx - 1]
                if header not in decision_headers:
                    continue

                cls = classify(cell.value)
                if cls == "pass":
                    cell.fill = fill_pass
                elif cls == "fail":
                    cell.fill = fill_fail
                elif cls == "unclear":
                    cell.fill = fill_unclear


def run_step6(config: dict) -> dict:
    out_dir = config.get("out_dir") or getattr(cfg, "out_dir", "outputs")
    viz_dir = os.path.join(out_dir, "step6")
    os.makedirs(viz_dir, exist_ok=True)

    # Prefer manual Step 5 output if present, otherwise fall back to default
    step5_csv_manual = os.path.join(out_dir, "step5", "step5_eligibility_wide_manual.csv")
    step5_csv_default = os.path.join(out_dir, "step5", "step5_eligibility_wide.csv")
    step5_csv = step5_csv_manual if os.path.exists(step5_csv_manual) else step5_csv_default
    step3_csv = os.path.join(out_dir, "step3", "step3_benchmark_list.csv")
    step4_csv = os.path.join(out_dir, "step4", "step4_abstracts.csv")

    out_excel = os.path.join(viz_dir, "eligibility_report.xlsx")

    if not os.path.exists(step5_csv):
        print("❌ Missing Step 5 output.")
        return {"status": "error"}

    print(f"[Step 6] Using Step 5 input: {step5_csv}")
    print(f"[Step 6] Generating visuals in: {viz_dir}")

    df_res = pd.read_csv(step5_csv)
    df_meta = pd.read_csv(step3_csv)

    abstract_map = {}
    if os.path.exists(step4_csv):
        print("   Found Step 4 Abstracts. Loading...")
        abstract_map = _load_abstract_map(step4_csv)
    else:
        print("   ⚠️ Step 4 file not found. Abstracts may be missing.")

    merged_base, crit_cols_base, cols_to_fix_base = _prepare_dataset(df_res, df_meta, abstract_map)

    _save_bar_summary(
        merged_base,
        cols_to_fix_base,
        viz_dir,
        "step6_1_criteria_summary.png",
        "Screening Results Breakdown",
    )

    df_inc = merged_base[merged_base["final_decision"].str.contains("include|yes|pass", case=False)].copy()
    _generate_long_heatmap(df_inc, crit_cols_base, viz_dir, "INCLUDED Papers", "step6_2_heatmap_included.png")

    df_exc = merged_base[merged_base["final_decision"].str.contains("exclude|no|fail", case=False)].copy()
    _generate_long_heatmap(df_exc, crit_cols_base, viz_dir, "EXCLUDED Papers", "step6_3_heatmap_excluded.png")

    df_unc = merged_base[~merged_base.index.isin(df_inc.index) & ~merged_base.index.isin(df_exc.index)].copy()
    _generate_long_heatmap(df_unc, crit_cols_base, viz_dir, "UNCLEAR Papers", "step6_4_heatmap_unclear.png")

    print("   📊 Generating Styled Excel...")
    _write_excel(merged_base, crit_cols_base, out_excel)
    print(f"   ✅ Saved Excel Report: {out_excel}")

    return {"status": "ok", "path": viz_dir}


if __name__ == "__main__":
    run_step6({})
